"""
可转债安全性评估 & 转股排名表格 - Excel 生成器
读取两个理杏仁导出的 CSV 文件，生成格式化的 Excel
"""

import pandas as pd
import os
import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式定义 ──
FILL_SAFE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_LOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_MEDIUM = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
FILL_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_NORMAL = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

RATING_FILL_MAP = {
    "安全": FILL_SAFE, "低风险": FILL_LOW,
    "中风险": FILL_MEDIUM, "高风险": FILL_HIGH,
}


def clean_value(val):
    """去除 Excel 公式前缀 '='，转为数值"""
    if isinstance(val, str):
        val = val.strip()
        if val.startswith("="):
            val = val[1:]
        # 尝试转数值
        try:
            return float(val)
        except:
            return val
    return val


def clean_dataframe(df):
    """清洗整个 DataFrame：去除 '=' 前缀，识别数值列"""
    for col in df.columns:
        df[col] = df[col].apply(clean_value)
    return df


def process_safety_csv(csv_path, output_path):
    """处理可转债安全性 CSV，生成带评级的 Excel"""
    print(f"\n{'='*60}")
    print("处理表格一：可转债安全性评估")
    print(f"{'='*60}")

    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    print(f"  读取 CSV: {len(df)} 行, {len(df.columns)} 列")

    # 列名
    cols = list(df.columns)
    print(f"  列名: {cols[:10]}...")

    # 清洗数据
    df = clean_dataframe(df)

    # 提取关键字段
    # 排除公式列（列名含运算引用），只匹配原始数据列
    def is_formula_col(name):
        """判断是否为计算列：含 [ 或明显公式引用"""
        return "[" in name and "]" in name

    col_map = {}
    for i, c in enumerate(cols):
        if is_formula_col(c):
            continue  # 跳过公式计算列
        if "代码" in c:
            col_map["代码"] = col_map.get("代码", i)
        elif "交易所" in c or "市场" in c:
            col_map["市场"] = col_map.get("市场", i)
        elif "公司" in c and col_map.get("公司") is None:
            col_map["公司"] = i
        elif "货币资金" in c:
            col_map["货币资金"] = col_map.get("货币资金", i)
        elif "交易性金融资产" in c:
            col_map["交易性金融资产"] = col_map.get("交易性金融资产", i)
        elif "有息负债" in c:
            col_map["有息负债"] = col_map.get("有息负债", i)
        elif "息税前利润" in c or "EBIT" in c:
            col_map["EBIT"] = col_map.get("EBIT", i)
        elif "市值" in c:
            col_map["市值"] = col_map.get("市值", i)
        elif "流动负债合计" in c:
            col_map["流动负债"] = col_map.get("流动负债", i)
        elif "负债合计" in c:
            col_map["总负债"] = col_map.get("总负债", i)
        elif "利息费用" in c or "利息支出" in c:
            col_map["利息支出"] = col_map.get("利息支出", i)
        elif "是否有可转债" in c:
            col_map["是否有可转债"] = col_map.get("是否有可转债", i)

    print(f"  字段映射: {col_map}")

    # 构建结果表
    result = pd.DataFrame()
    result["市场"] = df.iloc[:, col_map.get("市场", 0)]
    result["代码"] = df.iloc[:, col_map.get("代码", 1)]
    result["公司名称"] = df.iloc[:, col_map.get("公司", 2)]
    result["是否有可转债"] = df.iloc[:, col_map.get("是否有可转债", 10)]

    # 财务数据（单位：亿元）
    def fmt_yi(series):
        """将元转为亿元，保留2位小数"""
        return series.apply(lambda x: round(x / 1e8, 2) if isinstance(x, (int, float)) else x)

    result["货币资金(亿)"] = fmt_yi(df.iloc[:, col_map.get("货币资金", 11)])
    result["交易性金融资产(亿)"] = fmt_yi(df.iloc[:, col_map.get("交易性金融资产", 12)])
    result["流动负债(亿)"] = fmt_yi(df.iloc[:, col_map.get("流动负债", 17)])
    result["有息负债(亿)"] = fmt_yi(df.iloc[:, col_map.get("有息负债", 13)])
    result["总负债(亿)"] = fmt_yi(df.iloc[:, col_map.get("总负债", 16)])
    result["总市值(亿)"] = fmt_yi(df.iloc[:, col_map.get("市值", 15)])
    result["EBIT(亿)"] = fmt_yi(df.iloc[:, col_map.get("EBIT", 14)])
    result["利息支出(亿)"] = fmt_yi(df.iloc[:, col_map.get("利息支出", 8)])

    # ── 计算三个安全性指标 ──
    cash = df.iloc[:, col_map.get("货币资金", 11)]
    fin_assets = df.iloc[:, col_map.get("交易性金融资产", 12)]
    current_liab = df.iloc[:, col_map.get("流动负债", 17)]
    interest_debt = df.iloc[:, col_map.get("有息负债", 13)]
    total_liab = df.iloc[:, col_map.get("总负债", 16)]
    mkt_cap = df.iloc[:, col_map.get("市值", 15)]
    ebit = df.iloc[:, col_map.get("EBIT", 14)]
    interest_exp = df.iloc[:, col_map.get("利息支出", 8)]

    # 指标一：利息保障倍数 >= 7
    result["利息保障倍数"] = ebit / interest_exp
    result["指标一:利息保障倍数>=7"] = result["利息保障倍数"].apply(
        lambda x: "是" if isinstance(x, (int, float)) and x >= 7 else "否"
    )

    # 指标二：货币资金 + 交易性金融资产 >= 流动负债 或 有息负债
    cash_sum = cash + fin_assets
    cond2a = cash_sum >= current_liab
    cond2b = cash_sum >= interest_debt
    result["指标二:偿债能力"] = (cond2a | cond2b).apply(lambda x: "是" if x else "否")

    # 指标三：总负债 / 总市值 <= 1.5
    ratio3 = total_liab / mkt_cap
    result["负债市值比"] = ratio3
    result["指标三:负债/市值<=1.5"] = ratio3.apply(
        lambda x: "是" if isinstance(x, (int, float)) and x <= 1.5 else "否"
    )

    # 符合条件数
    result["符合条件数"] = result.apply(
        lambda row: sum([
            row["指标一:利息保障倍数>=7"] == "是",
            row["指标二:偿债能力"] == "是",
            row["指标三:负债/市值<=1.5"] == "是",
        ]), axis=1
    )

    # 安全性评级
    safety_map = {3: "安全", 2: "低风险", 1: "中风险", 0: "高风险"}
    result["安全性评级"] = result["符合条件数"].map(safety_map)

    # 只保留有可转债的公司
    has_cb = df.iloc[:, col_map.get("是否有可转债", 10)] == 1
    result = result[has_cb].reset_index(drop=True)

    # 统计数据
    print(f"\n  有可转债的公司: {len(result)} 家")
    print(f"  评级分布:")
    for rating in ["安全", "低风险", "中风险", "高风险"]:
        cnt = (result["安全性评级"] == rating).sum()
        print(f"    {rating}: {cnt} 家")

    # 输出列顺序
    output_cols = [
        "市场", "代码", "公司名称", "是否有可转债",
        "货币资金(亿)", "交易性金融资产(亿)", "流动负债(亿)", "有息负债(亿)",
        "总负债(亿)", "总市值(亿)", "EBIT(亿)", "利息支出(亿)",
        "利息保障倍数", "负债市值比",
        "指标一:利息保障倍数>=7", "指标二:偿债能力", "指标三:负债/市值<=1.5",
        "符合条件数", "安全性评级"
    ]

    result = result[output_cols]

    # 生成 Excel
    result.to_excel(output_path, index=False, sheet_name="可转债安全性")
    print(f"\n  Excel 已保存: {output_path}")

    # ── 格式化 ──
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "可转债安全性"

    # 表头
    for col_idx in range(1, len(result.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # 数据行
    rating_col_idx = list(result.columns).index("安全性评级") + 1
    for row_idx in range(2, len(result) + 2):
        for col_idx in range(1, len(result.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        rating_val = ws.cell(row=row_idx, column=rating_col_idx).value
        if rating_val in RATING_FILL_MAP:
            fill = RATING_FILL_MAP[rating_val]
            for col_idx in range(1, len(result.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # 列宽
    for col_idx in range(1, len(result.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(result) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(result.columns))}{len(result) + 1}"
    wb.save(output_path)
    print(f"  [OK] 格式完成")


def process_bond_csv(csv_path, output_path):
    """处理转股排名表格 CSV，生成格式化 Excel"""
    print(f"\n{'='*60}")
    print("处理表格二：转股排名表格")
    print(f"{'='*60}")

    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    print(f"  读取 CSV: {len(df)} 行, {len(df.columns)} 列")

    df = clean_dataframe(df)
    cols = list(df.columns)
    print(f"  列名: {cols}")

    # 输出 Excel
    df.to_excel(output_path, index=False, sheet_name="转股排名")
    print(f"\n  Excel 已保存: {output_path}")

    # ── 格式化 ──
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "转股排名"

    # 表头
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # 数据行
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(df) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 25)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    wb.save(output_path)
    print(f"  ✅ 格式化完成！")


def main():
    base = r"C:\Users\daicunzai\WorkBuddy\2026-06-14-13-30-54"
    exports = os.path.join(base, "exports")
    output_dir = base

    csv1 = os.path.join(exports, "可转债安全性.csv")
    csv2 = os.path.join(exports, "转股排名表格.csv")

    # 表格一
    if os.path.exists(csv1):
        out1 = os.path.join(output_dir, "可转债安全性评估.xlsx")
        process_safety_csv(csv1, out1)
    else:
        print(f"❌ 未找到: {csv1}")

    # 表格二
    if os.path.exists(csv2):
        out2 = os.path.join(output_dir, "转股排名表格.xlsx")
        process_bond_csv(csv2, out2)
    else:
        print(f"❌ 未找到: {csv2}")

    print(f"\n{'='*60}")
    print("✅ 全部完成！生成的文件:")
    for f in ["可转债安全性评估.xlsx", "转股排名表格.xlsx"]:
        fp = os.path.join(output_dir, f)
        if os.path.exists(fp):
            print(f"  📄 {fp} ({os.path.getsize(fp)/1024:.1f} KB)")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
